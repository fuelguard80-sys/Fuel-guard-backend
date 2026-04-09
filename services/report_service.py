from __future__ import annotations

import csv
import io
import logging
from datetime import datetime, timezone
from typing import Any

from core.firebase import Collections

logger = logging.getLogger(__name__)


# ── Firestore data helpers ────────────────────────────────────────────────────

def build_transaction_report(
    db,
    start: datetime,
    end: datetime,
    station_id: str | None = None,
    fuel_type: str | None = None,
) -> dict[str, Any]:
    query = db.collection(Collections.TRANSACTIONS).where("created_at", ">=", start).where(
        "created_at", "<=", end
    )
    if station_id:
        query = query.where("station_id", "==", station_id)
    if fuel_type:
        query = query.where("fuel_type", "==", fuel_type)

    docs = query.get()
    items = [{"id": d.id, **d.to_dict()} for d in docs]
    total_revenue = sum(t.get("total_amount", 0) for t in items)
    total_litres  = sum(t.get("litres_dispensed", 0) for t in items)

    return {
        "period": {"from": start.isoformat(), "to": end.isoformat()},
        "total_transactions": len(items),
        "total_revenue_pkr": round(total_revenue, 2),
        "total_litres": round(total_litres, 2),
        "items": items,
    }


def build_sales_report(
    db,
    start: datetime,
    end: datetime,
    station_id: str | None = None,
) -> dict[str, Any]:
    query = db.collection(Collections.TRANSACTIONS).where("created_at", ">=", start).where(
        "created_at", "<=", end
    )
    if station_id:
        query = query.where("station_id", "==", station_id)

    docs  = query.get()
    items = [d.to_dict() for d in docs]

    by_fuel: dict[str, dict] = {}
    by_payment: dict[str, int] = {}

    for t in items:
        ft = t.get("fuel_type", "unknown")
        pm = t.get("payment_method", "unknown")

        if ft not in by_fuel:
            by_fuel[ft] = {"count": 0, "revenue": 0.0, "litres": 0.0}
        by_fuel[ft]["count"]   += 1
        by_fuel[ft]["revenue"] += t.get("total_amount", 0)
        by_fuel[ft]["litres"]  += t.get("litres_dispensed", 0)

        by_payment[pm] = by_payment.get(pm, 0) + 1

    return {
        "period": {"from": start.isoformat(), "to": end.isoformat()},
        "total_revenue_pkr": round(sum(t.get("total_amount", 0) for t in items), 2),
        "by_fuel_type": by_fuel,
        "by_payment_method": by_payment,
    }


# ── Receipt PDF ───────────────────────────────────────────────────────────────

def generate_receipt_pdf(transaction: dict) -> bytes:
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A5
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Spacer

    def _fmt_datetime(value) -> str:
        if value is None:
            return "N/A"
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d %H:%M:%S UTC")
        return str(value)[:19]

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A5,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )
    styles   = getSampleStyleSheet()
    centered = ParagraphStyle("centered", parent=styles["Normal"], alignment=1)
    story    = [
        Paragraph("FUEL GUARD", ParagraphStyle("title", parent=styles["Heading1"], alignment=1)),
        Paragraph("Transaction Receipt", centered),
        Spacer(1, 10),
        HRFlowable(width="100%", color=colors.grey),
        Spacer(1, 6),
    ]

    tx_id = transaction.get("id", "N/A")
    short_id = f"...{tx_id[-8:]}" if len(tx_id) > 8 else tx_id

    fields = [
        ("Transaction",     short_id),
        ("Date",            _fmt_datetime(transaction.get("created_at"))),
        ("Fuel Type",       transaction.get("fuel_type", "N/A").title()),
        ("Litres Dispensed", f"{transaction.get('litres_dispensed', 0):.2f} L"),
        ("Price / Litre",   f"PKR {transaction.get('price_per_litre', 0):.2f}"),
        ("Total Amount",    f"PKR {transaction.get('total_amount', 0):,.2f}"),
        ("Payment Method",  transaction.get("payment_method", "N/A").replace("_", " ").title()),
        ("Status",          transaction.get("status", "N/A").title()),
    ]

    for label, value in fields:
        story.append(Paragraph(f"<b>{label}:</b> {value}", styles["Normal"]))
        story.append(Spacer(1, 4))

    story += [
        Spacer(1, 10),
        HRFlowable(width="100%", color=colors.grey),
        Spacer(1, 6),
        Paragraph("Thank you for using Fuel Guard.", centered),
    ]

    doc.build(story)
    return buf.getvalue()


# ── Report export ─────────────────────────────────────────────────────────────

def export_report_to_pdf(data: dict, report_type: str) -> bytes:
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf   = BytesIO()
    doc   = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15 * mm, rightMargin=15 * mm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"Fuel Guard — {report_type.replace('_', ' ').title()} Report", styles["Heading1"]),
        Paragraph(
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC",
            styles["Normal"],
        ),
        Spacer(1, 12),
    ]

    items = data.get("items", [])
    if items:
        headers = list(items[0].keys())[:8]
        rows    = [headers] + [
            [str(item.get(h, ""))[:30] for h in headers] for item in items[:100]
        ]
        table = Table(rows)
        table.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0),  colors.HexColor("#1C2536")),
            ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
            ("FONTSIZE",     (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ("GRID",         (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("ALIGN",        (0, 0), (-1, -1), "LEFT"),
        ]))
        story.append(table)
    else:
        story.append(Paragraph("No data available for this period.", styles["Normal"]))

    doc.build(story)
    return buf.getvalue()


def export_report_to_excel(data: dict, report_type: str) -> bytes:
    from io import BytesIO

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type.replace("_", " ").title()

    header_fill = PatternFill("solid", fgColor="1C2536")
    header_font = Font(color="FFFFFF", bold=True)

    items = data.get("items", [])
    if not items:
        ws.append(["No data available for this period."])
    else:
        headers = list(items[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center")
        for item in items:
            ws.append([str(item.get(h, "")) for h in headers])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_report_to_csv(data: dict, report_type: str) -> bytes:
    items = data.get("items", [])
    if not items:
        return b"No data available for this period."

    headers   = list(items[0].keys())
    text_buf  = io.StringIO()
    writer    = csv.DictWriter(text_buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for item in items:
        writer.writerow({k: str(v) for k, v in item.items()})

    return text_buf.getvalue().encode("utf-8")
